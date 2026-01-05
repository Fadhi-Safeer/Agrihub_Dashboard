import os
import shutil
import uuid
from datetime import datetime, timedelta
import cv2
import numpy as np
from classification_mapper import ClassificationMapper
from pathlib import Path
from datetime import datetime
import json
import threading
from openpyxl import Workbook, load_workbook
from typing import Optional



# Track last save time per camera
_last_saved_at: dict[str, datetime] = {}


def get_timestamp_paths(base_dir):
    now = datetime.now()
    month_folder = now.strftime("%B_%Y")       # e.g., July_2025
    date_folder = now.strftime("%Y_%m_%d")     # e.g., 2025_07_09
    time_folder = now.strftime("%H_%M")  # e.g., 14_04 for 2:04 PM
    full_path = os.path.join(base_dir, month_folder, date_folder, time_folder)
    return full_path, now


def save_frame_locally(
    frame: np.ndarray,
    cam_num: str,
    classification_results: dict,
    base_dir: str = "C:/Users/Fadhi Safeer/OneDrive/Documents/Internship/Agri hub/STORAGE/camera_storage"
) -> None:
    """
    Saves classified crop frame in a timestamped camera folder with structured naming.
    """
    print(f"[DEBUG] Started Saving classified frame for camera {cam_num}...")

    # Check time restriction
    now = datetime.now()

    print(f"[DEBUG] Surpassed Saving classified frame for camera {cam_num}...")
    # Get time-based directory structure
    save_path, _ = get_timestamp_paths(base_dir)

    # Create camera-specific directory
    cam_dir = os.path.join(save_path, str(cam_num))
    os.makedirs(cam_dir, exist_ok=True)
    
    # ✅ Check how many images already exist
    existing_images = [f for f in os.listdir(cam_dir) if f.endswith('.jpg')]
    print(existing_images)
    if len(existing_images) >= 7:
        print(f"[SKIPPED] Max image count (7) reached for {cam_num} in {cam_dir}")
        return

    # Normalize classification fields
    growth_stage = str(classification_results.get("growth", "unknown")).lower().strip().replace(" ", "_")
    health_status = str(classification_results.get("health", "unknown")).lower().strip().replace(" ", "_")
    disease_type = str(classification_results.get("disease", "")).lower().strip().replace(" ", "_")

    # Convert to short codes
    growth_code = ClassificationMapper.get_growth_code(growth_stage)
    health_code = ClassificationMapper.get_health_code(health_status)
    disease_code = ClassificationMapper.get_disease_code(disease_type)

    # Filename format
    filename = f"{cam_num}_{growth_code}_{health_code}_{disease_code}_{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.jpg"

    # Save
    cv2.imwrite(os.path.join(cam_dir, filename), frame)
    _last_saved_at[cam_num] = now


    print(f"[SAVED] Classified image: {filename}")


def save_camera_view_frame(
    frame: np.ndarray,
    cam_num: str,
    base_dir: str = "C:/Users/Fadhi Safeer/OneDrive/Documents/Internship/Agri hub/STORAGE/camera_storage"
) -> None:
    """
    Saves the latest camera snapshot in a timestamped camera folder.
    """

    # Check time restriction
    now = datetime.now()


    # Get time-based directory structure
    save_path, _ = get_timestamp_paths(base_dir)

    # Camera view folder
    cam_view_dir = os.path.join(save_path, str(cam_num), "camera_view")
    os.makedirs(cam_view_dir, exist_ok=True)

    # Clean old snapshot(s)
    clear_directory(cam_view_dir)

    # Filename
    filename = f"{cam_num}_{now.strftime('%Y%m%d_%H%M%S')}.jpg"

    # Save
    cv2.imwrite(os.path.join(cam_view_dir, filename), frame)
    #print(f"[SAVED] Snapshot: {os.path.join(cam_view_dir, filename)}")


def clear_directory(directory_path):
    """
    Clears all files and subdirectories from the given path.
    """
    if not os.path.exists(directory_path):
        #print(f"[WARN] Directory '{directory_path}' does not exist.")
        return 0

    if not os.path.isdir(directory_path):
        #print(f"[ERROR] '{directory_path}' is not a directory.")
        return 0

    deleted_count = 0
    try:
        for item in os.listdir(directory_path):
            item_path = os.path.join(directory_path, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
                deleted_count += 1
            elif os.path.isdir(item_path):
                shutil.rmtree(item_path)
                deleted_count += 1
        return deleted_count
    except Exception as e:
        #print(f"[ERROR] Failed to clear directory: {e}")
        return deleted_count



_EXCEL_LOCK = threading.Lock()

HEADERS = [
    "date", "time", "camera_number", "plant_id",
    "growth", "health", "disease",
    "disease_status", "health_status",
]

DATA_JSON_PATH = "backend/Data/Data.json"


def _get_agrivision_excel_path() -> Path:
    data_path = Path(DATA_JSON_PATH)
    if not data_path.exists():
        raise FileNotFoundError(f"Data.json not found: {data_path}")

    data = json.loads(data_path.read_text(encoding="utf-8"))

    p = data.get("agrivison_data_path")
    if not isinstance(p, str) or not p.strip():
        raise ValueError("Data.json missing or invalid 'agrivison_data_path'")

    return Path(p)


def append_agrivision_row(
    *,
    camera_number: str,
    plant_id: str,
    growth: str,
    health: str,
    disease: str,
    disease_status: Optional[int] = None,  # 0/1/None
    health_status: Optional[int] = None,   # 0/1/None
) -> str:
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    excel_path = _get_agrivision_excel_path()
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    with _EXCEL_LOCK:
        if excel_path.exists():
            wb = load_workbook(excel_path)
            ws = wb.active

            # Ensure headers exist and match
            if ws.max_row == 0:
                ws.append(HEADERS)
            else:
                first_row = [c.value for c in ws[1]]
                if first_row != HEADERS:
                    # overwrite header row
                    for col, h in enumerate(HEADERS, start=1):
                        ws.cell(row=1, column=col, value=h)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "agrivision"
            ws.append(HEADERS)

        ws.append([
            date_str, time_str, camera_number, plant_id,
            growth, health, disease,
            "" if disease_status is None else disease_status,
            "" if health_status is None else health_status,
        ])

        wb.save(excel_path)

    return str(excel_path)